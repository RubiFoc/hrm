"""Integration tests for accountant workspace APIs."""

from __future__ import annotations

from datetime import UTC, datetime, timedelta
from io import BytesIO
from pathlib import Path
from uuid import UUID, uuid4

import pytest
from httpx import ASGITransport, AsyncClient
from openpyxl import load_workbook
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session

from hrm_backend.audit.models.event import AuditEvent
from hrm_backend.auth.dependencies.auth import get_current_auth_context
from hrm_backend.auth.schemas.token_claims import AuthContext
from hrm_backend.core.models.base import Base
from hrm_backend.employee.models.onboarding import OnboardingRun, OnboardingTask
from hrm_backend.employee.models.profile import EmployeeProfile
from hrm_backend.main import app
from hrm_backend.settings import AppSettings, get_settings

pytestmark = pytest.mark.anyio


@pytest.fixture()
def sqlite_database_url(tmp_path: Path) -> str:
    """Provide temporary SQLite URL for accountant workspace integration tests."""
    return f"sqlite+pysqlite:///{tmp_path / 'accounting_workspace_api.db'}"


@pytest.fixture()
def configured_app(sqlite_database_url: str):
    """Configure app dependency overrides for accountant workspace integration tests."""
    settings = AppSettings(
        database_url=sqlite_database_url,
        redis_url="redis://localhost:6379/15",
        jwt_secret="integration-secret-with-minimum-32-bytes",
    )
    context_holder = {
        "context": AuthContext(
            subject_id=uuid4(),
            role="accountant",
            session_id=uuid4(),
            token_id=uuid4(),
            expires_at=9999999999,
        )
    }

    def _get_settings_override() -> AppSettings:
        return settings

    def _get_auth_context_override() -> AuthContext:
        return context_holder["context"]

    app.dependency_overrides[get_settings] = _get_settings_override
    app.dependency_overrides[get_current_auth_context] = _get_auth_context_override

    engine = create_engine(sqlite_database_url, future=True)
    Base.metadata.create_all(engine)
    seeded = _seed_accounting_records(sqlite_database_url)
    try:
        yield app, context_holder, sqlite_database_url, seeded
    finally:
        app.dependency_overrides.pop(get_settings, None)
        app.dependency_overrides.pop(get_current_auth_context, None)
        engine.dispose()


def _seed_accounting_records(database_url: str) -> dict[str, str]:
    """Insert employee profiles, onboarding runs, and tasks for accountant API tests."""
    accountant_subject_id = "aaaaaaaa-aaaa-4aaa-8aaa-aaaaaaaaaaaa"
    engine = create_engine(database_url, future=True)
    try:
        with Session(engine) as session:
            session.add_all(
                [
                    EmployeeProfile(
                        employee_id="11111111-1111-4111-8111-111111111111",
                        hire_conversion_id="21111111-1111-4111-8111-111111111111",
                        vacancy_id="31111111-1111-4111-8111-111111111111",
                        candidate_id="41111111-1111-4111-8111-111111111111",
                        first_name="Ada",
                        last_name="Adams",
                        email="ada@example.com",
                        phone=None,
                        location="Minsk",
                        current_title="Accountant Liaison",
                        extra_data_json={},
                        offer_terms_summary="Payroll starter pack",
                        start_date=datetime(2026, 4, 1, tzinfo=UTC).date(),
                        created_by_staff_id="51111111-1111-4111-8111-111111111111",
                    ),
                    EmployeeProfile(
                        employee_id="12222222-2222-4222-8222-222222222222",
                        hire_conversion_id="22222222-2222-4222-8222-222222222222",
                        vacancy_id="32222222-2222-4222-8222-222222222222",
                        candidate_id="42222222-2222-4222-8222-222222222222",
                        first_name="Grace",
                        last_name="Baker",
                        email="grace@example.com",
                        phone=None,
                        location="Brest",
                        current_title="Engineer",
                        extra_data_json={},
                        offer_terms_summary="Benefit enrollment",
                        start_date=datetime(2026, 4, 15, tzinfo=UTC).date(),
                        created_by_staff_id="52222222-2222-4222-8222-222222222222",
                    ),
                    EmployeeProfile(
                        employee_id="13333333-3333-4333-8333-333333333333",
                        hire_conversion_id="23333333-3333-4333-8333-333333333333",
                        vacancy_id="33333333-3333-4333-8333-333333333333",
                        candidate_id="43333333-3333-4333-8333-333333333333",
                        first_name="Tim",
                        last_name="Clark",
                        email="tim@example.com",
                        phone=None,
                        location="Grodno",
                        current_title="Designer",
                        extra_data_json={},
                        offer_terms_summary="Design onboarding",
                        start_date=datetime(2026, 4, 20, tzinfo=UTC).date(),
                        created_by_staff_id="53333333-3333-4333-8333-333333333333",
                    ),
                ]
            )
            session.add_all(
                [
                    OnboardingRun(
                        onboarding_id="61111111-1111-4111-8111-111111111111",
                        employee_id="11111111-1111-4111-8111-111111111111",
                        hire_conversion_id="21111111-1111-4111-8111-111111111111",
                        status="started",
                        started_at=datetime(2026, 3, 12, 9, 0, tzinfo=UTC),
                        started_by_staff_id="71111111-1111-4111-8111-111111111111",
                    ),
                    OnboardingRun(
                        onboarding_id="62222222-2222-4222-8222-222222222222",
                        employee_id="12222222-2222-4222-8222-222222222222",
                        hire_conversion_id="22222222-2222-4222-8222-222222222222",
                        status="started",
                        started_at=datetime(2026, 3, 11, 9, 0, tzinfo=UTC),
                        started_by_staff_id="72222222-2222-4222-8222-222222222222",
                    ),
                    OnboardingRun(
                        onboarding_id="63333333-3333-4333-8333-333333333333",
                        employee_id="13333333-3333-4333-8333-333333333333",
                        hire_conversion_id="23333333-3333-4333-8333-333333333333",
                        status="started",
                        started_at=datetime(2026, 3, 10, 9, 0, tzinfo=UTC),
                        started_by_staff_id="73333333-3333-4333-8333-333333333333",
                    ),
                ]
            )
            session.add_all(
                [
                    OnboardingTask(
                        task_id="81111111-1111-4111-8111-111111111111",
                        onboarding_id="61111111-1111-4111-8111-111111111111",
                        template_id="91111111-1111-4111-8111-111111111111",
                        template_item_id="a1111111-1111-4111-8111-111111111111",
                        code="collect_bank_details",
                        title="Collect bank details",
                        description=None,
                        sort_order=10,
                        is_required=True,
                        status="pending",
                        assigned_role="accountant",
                        assigned_staff_id=None,
                        due_at=datetime.now(UTC) - timedelta(days=1),
                        completed_at=None,
                    ),
                    OnboardingTask(
                        task_id="82222222-2222-4222-8222-222222222222",
                        onboarding_id="61111111-1111-4111-8111-111111111111",
                        template_id="92222222-2222-4222-8222-222222222222",
                        template_item_id="a2222222-2222-4222-8222-222222222222",
                        code="benefit_setup",
                        title="Benefit setup",
                        description=None,
                        sort_order=20,
                        is_required=True,
                        status="completed",
                        assigned_role="accountant",
                        assigned_staff_id=None,
                        due_at=datetime.now(UTC) + timedelta(days=2),
                        completed_at=datetime.now(UTC) - timedelta(hours=2),
                    ),
                    OnboardingTask(
                        task_id="83333333-3333-4333-8333-333333333333",
                        onboarding_id="62222222-2222-4222-8222-222222222222",
                        template_id="93333333-3333-4333-8333-333333333333",
                        template_item_id="a3333333-3333-4333-8333-333333333333",
                        code="tax_form",
                        title="Tax form",
                        description=None,
                        sort_order=10,
                        is_required=True,
                        status="in_progress",
                        assigned_role="hr",
                        assigned_staff_id=accountant_subject_id,
                        due_at=datetime.now(UTC) + timedelta(days=3),
                        completed_at=None,
                    ),
                    OnboardingTask(
                        task_id="84444444-4444-4444-8444-444444444444",
                        onboarding_id="63333333-3333-4333-8333-333333333333",
                        template_id="94444444-4444-4444-8444-444444444444",
                        template_item_id="a4444444-4444-4444-8444-444444444444",
                        code="manager_intro",
                        title="Manager intro",
                        description=None,
                        sort_order=10,
                        is_required=True,
                        status="pending",
                        assigned_role="manager",
                        assigned_staff_id=None,
                        due_at=datetime.now(UTC) + timedelta(days=1),
                        completed_at=None,
                    ),
                ]
            )
            session.commit()
    finally:
        engine.dispose()
    return {"accountant_subject_id": accountant_subject_id}


def _load_events(database_url: str) -> list[AuditEvent]:
    """Load ordered audit events for assertions."""
    engine = create_engine(database_url, future=True)
    try:
        with Session(engine) as session:
            return list(
                session.execute(
                    select(AuditEvent).order_by(AuditEvent.occurred_at, AuditEvent.event_id)
                ).scalars()
            )
    finally:
        engine.dispose()


@pytest.fixture()
async def api_client(configured_app) -> AsyncClient:
    """Provide async API client for accountant workspace integration tests."""
    configured, *_ = configured_app
    async with AsyncClient(
        transport=ASGITransport(app=configured),
        base_url="http://testserver",
    ) as client:
        yield client


async def test_accounting_workspace_api_lists_and_exports_visible_scope(
    configured_app,
    api_client: AsyncClient,
) -> None:
    """Verify accountant workspace list and dual-format exports use the same filtered scope."""
    _, context_holder, database_url, seeded = configured_app
    context_holder["context"] = AuthContext(
        subject_id=UUID(seeded["accountant_subject_id"]),
        role="accountant",
        session_id=uuid4(),
        token_id=uuid4(),
        expires_at=9999999999,
    )

    list_response = await api_client.get(
        "/api/v1/accounting/workspace",
        params={"limit": 1, "offset": 0},
    )
    assert list_response.status_code == 200
    list_payload = list_response.json()
    assert list_payload["total"] == 2
    assert len(list_payload["items"]) == 1
    assert list_payload["items"][0]["last_name"] == "Adams"

    csv_response = await api_client.get(
        "/api/v1/accounting/workspace/export",
        params={"format": "csv"},
    )
    assert csv_response.status_code == 200
    assert csv_response.headers["content-type"].startswith("text/csv")
    assert "attachment; filename=\"accounting-workspace-" in (
        csv_response.headers["content-disposition"]
    )
    csv_lines = csv_response.text.strip().splitlines()
    assert len(csv_lines) == 3
    assert all("Clark" not in line for line in csv_lines)

    xlsx_response = await api_client.get(
        "/api/v1/accounting/workspace/export",
        params={"format": "xlsx", "search": "payroll"},
    )
    assert xlsx_response.status_code == 200
    workbook = load_workbook(filename=BytesIO(xlsx_response.content))
    worksheet = workbook["accounting_workspace"]
    values = list(worksheet.iter_rows(values_only=True))
    assert len(values) == 2
    assert values[1][2] == "Ada"
    assert values[1][3] == "Adams"

    empty_csv = await api_client.get(
        "/api/v1/accounting/workspace/export",
        params={"format": "csv", "search": "no-match"},
    )
    assert empty_csv.status_code == 200
    assert len(empty_csv.text.strip().splitlines()) == 1

    events = _load_events(database_url)
    success_actions = [event.action for event in events if event.result == "success"]
    assert "accounting_workspace:read" in success_actions
    assert "accounting_export:download" in success_actions


@pytest.mark.parametrize("role", ["hr", "manager", "leader", "employee"])
async def test_accounting_workspace_api_denies_non_accounting_roles(
    configured_app,
    api_client: AsyncClient,
    role: str,
) -> None:
    """Verify non-accountant roles are denied by the accountant workspace RBAC boundary."""
    _, context_holder, database_url, seeded = configured_app
    context_holder["context"] = AuthContext(
        subject_id=UUID(seeded["accountant_subject_id"]),
        role=role,
        session_id=uuid4(),
        token_id=uuid4(),
        expires_at=9999999999,
    )

    response = await api_client.get("/api/v1/accounting/workspace")
    assert response.status_code == 403

    events = _load_events(database_url)
    denied_actions = [(event.action, event.result, event.actor_role) for event in events]
    assert ("accounting:read", "denied", role) in denied_actions


async def test_accounting_workspace_api_allows_admin_boundary(
    configured_app,
    api_client: AsyncClient,
) -> None:
    """Verify admin actors can call accountant workspace APIs directly."""
    _, context_holder, _, seeded = configured_app
    context_holder["context"] = AuthContext(
        subject_id=UUID(seeded["accountant_subject_id"]),
        role="admin",
        session_id=uuid4(),
        token_id=uuid4(),
        expires_at=9999999999,
    )

    response = await api_client.get(
        "/api/v1/accounting/workspace/export",
        params={"format": "csv", "search": "grace"},
    )
    assert response.status_code == 200
    assert "Grace" in response.text
