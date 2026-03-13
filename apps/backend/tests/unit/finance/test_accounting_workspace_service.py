"""Unit tests for accountant workspace visibility and export helpers."""

from __future__ import annotations

from datetime import UTC, datetime, timedelta
from io import BytesIO
from uuid import UUID, uuid4

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session
from starlette.requests import Request

from hrm_backend.audit.services.audit_service import AuditService
from hrm_backend.auth.schemas.token_claims import AuthContext
from hrm_backend.core.models.base import Base
from hrm_backend.employee.dao.onboarding_task_dao import OnboardingTaskDAO
from hrm_backend.employee.models.onboarding import OnboardingRun, OnboardingTask
from hrm_backend.employee.models.profile import EmployeeProfile
from hrm_backend.finance.dao import AccountingWorkspaceDAO
from hrm_backend.finance.services.accounting_workspace_service import AccountingWorkspaceService
from hrm_backend.finance.utils.exports import ACCOUNTING_WORKSPACE_EXPORT_COLUMNS


class _AuditServiceStub(AuditService):
    """Audit double that records accountant workspace events in memory."""

    def __init__(self) -> None:
        self.events: list[dict[str, object]] = []

    def record_api_event(self, **kwargs) -> None:  # type: ignore[override]
        """Capture API audit payloads for focused service assertions."""
        self.events.append(kwargs)


def _build_request(path: str) -> Request:
    """Create minimal request object for service methods."""
    return Request(
        {
            "type": "http",
            "method": "GET",
            "path": path,
            "headers": [],
            "client": ("127.0.0.1", 8000),
        }
    )


def _build_auth_context(*, role: str, subject_id: str) -> AuthContext:
    """Create deterministic auth context for accountant workspace tests."""
    return AuthContext(
        subject_id=UUID(subject_id),
        role=role,
        session_id=uuid4(),
        token_id=uuid4(),
        expires_at=9999999999,
    )


def _seed_accounting_workspace(session: Session) -> dict[str, str]:
    """Insert employee, onboarding, and task fixtures for accountant visibility tests."""
    accountant_subject_id = "aaaaaaaa-aaaa-4aaa-8aaa-aaaaaaaaaaaa"
    session.add_all(
        [
            EmployeeProfile(
                employee_id="11111111-1111-4111-8111-111111111111",
                hire_conversion_id="21111111-1111-4111-8111-111111111111",
                vacancy_id="31111111-1111-4111-8111-111111111111",
                candidate_id="41111111-1111-4111-8111-111111111111",
                first_name="Ada",
                last_name="Adams",
                email="ada@example.com",
                phone=None,
                location="Minsk",
                current_title="Accounting Analyst",
                extra_data_json={},
                offer_terms_summary="Payroll starter pack",
                start_date=datetime(2026, 4, 1, tzinfo=UTC).date(),
                created_by_staff_id="51111111-1111-4111-8111-111111111111",
            ),
            EmployeeProfile(
                employee_id="12222222-2222-4222-8222-222222222222",
                hire_conversion_id="22222222-2222-4222-8222-222222222222",
                vacancy_id="32222222-2222-4222-8222-222222222222",
                candidate_id="42222222-2222-4222-8222-222222222222",
                first_name="Grace",
                last_name="Baker",
                email="grace@example.com",
                phone=None,
                location="Brest",
                current_title="Engineer",
                extra_data_json={},
                offer_terms_summary="Benefit enrollment",
                start_date=datetime(2026, 4, 15, tzinfo=UTC).date(),
                created_by_staff_id="52222222-2222-4222-8222-222222222222",
            ),
            EmployeeProfile(
                employee_id="13333333-3333-4333-8333-333333333333",
                hire_conversion_id="23333333-3333-4333-8333-333333333333",
                vacancy_id="33333333-3333-4333-8333-333333333333",
                candidate_id="43333333-3333-4333-8333-333333333333",
                first_name="Tim",
                last_name="Clark",
                email="tim@example.com",
                phone=None,
                location="Grodno",
                current_title="Designer",
                extra_data_json={},
                offer_terms_summary="Design onboarding",
                start_date=datetime(2026, 4, 20, tzinfo=UTC).date(),
                created_by_staff_id="53333333-3333-4333-8333-333333333333",
            ),
        ]
    )
    session.add_all(
        [
            OnboardingRun(
                onboarding_id="61111111-1111-4111-8111-111111111111",
                employee_id="11111111-1111-4111-8111-111111111111",
                hire_conversion_id="21111111-1111-4111-8111-111111111111",
                status="started",
                started_at=datetime(2026, 3, 12, 9, 0, tzinfo=UTC),
                started_by_staff_id="71111111-1111-4111-8111-111111111111",
            ),
            OnboardingRun(
                onboarding_id="62222222-2222-4222-8222-222222222222",
                employee_id="12222222-2222-4222-8222-222222222222",
                hire_conversion_id="22222222-2222-4222-8222-222222222222",
                status="started",
                started_at=datetime(2026, 3, 11, 9, 0, tzinfo=UTC),
                started_by_staff_id="72222222-2222-4222-8222-222222222222",
            ),
            OnboardingRun(
                onboarding_id="63333333-3333-4333-8333-333333333333",
                employee_id="13333333-3333-4333-8333-333333333333",
                hire_conversion_id="23333333-3333-4333-8333-333333333333",
                status="started",
                started_at=datetime(2026, 3, 10, 9, 0, tzinfo=UTC),
                started_by_staff_id="73333333-3333-4333-8333-333333333333",
            ),
        ]
    )
    session.add_all(
        [
            OnboardingTask(
                task_id="81111111-1111-4111-8111-111111111111",
                onboarding_id="61111111-1111-4111-8111-111111111111",
                template_id="91111111-1111-4111-8111-111111111111",
                template_item_id="a1111111-1111-4111-8111-111111111111",
                code="collect_bank_details",
                title="Collect bank details",
                description=None,
                sort_order=10,
                is_required=True,
                status="pending",
                assigned_role="accountant",
                assigned_staff_id=None,
                due_at=datetime.now(UTC) - timedelta(days=1),
                completed_at=None,
            ),
            OnboardingTask(
                task_id="82222222-2222-4222-8222-222222222222",
                onboarding_id="61111111-1111-4111-8111-111111111111",
                template_id="92222222-2222-4222-8222-222222222222",
                template_item_id="a2222222-2222-4222-8222-222222222222",
                code="benefit_setup",
                title="Benefit setup",
                description=None,
                sort_order=20,
                is_required=True,
                status="completed",
                assigned_role="accountant",
                assigned_staff_id=None,
                due_at=datetime.now(UTC) + timedelta(days=2),
                completed_at=datetime.now(UTC) - timedelta(hours=2),
            ),
            OnboardingTask(
                task_id="83333333-3333-4333-8333-333333333333",
                onboarding_id="62222222-2222-4222-8222-222222222222",
                template_id="93333333-3333-4333-8333-333333333333",
                template_item_id="a3333333-3333-4333-8333-333333333333",
                code="tax_form",
                title="Tax form",
                description=None,
                sort_order=10,
                is_required=True,
                status="in_progress",
                assigned_role="hr",
                assigned_staff_id=accountant_subject_id,
                due_at=datetime.now(UTC) + timedelta(days=3),
                completed_at=None,
            ),
            OnboardingTask(
                task_id="84444444-4444-4444-8444-444444444444",
                onboarding_id="63333333-3333-4333-8333-333333333333",
                template_id="94444444-4444-4444-8444-444444444444",
                template_item_id="a4444444-4444-4444-8444-444444444444",
                code="manager_intro",
                title="Manager intro",
                description=None,
                sort_order=10,
                is_required=True,
                status="pending",
                assigned_role="manager",
                assigned_staff_id=None,
                due_at=datetime.now(UTC) + timedelta(days=1),
                completed_at=None,
            ),
        ]
    )
    session.commit()
    return {"accountant_subject_id": accountant_subject_id}


def test_accounting_workspace_list_is_scoped_and_ordered() -> None:
    """Verify accountant workspace includes only accountant-visible runs in deterministic order."""
    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    Base.metadata.create_all(engine)
    audit_service = _AuditServiceStub()
    with Session(engine) as session:
        seeded = _seed_accounting_workspace(session)
        service = AccountingWorkspaceService(
            workspace_dao=AccountingWorkspaceDAO(session=session),
            task_dao=OnboardingTaskDAO(session=session),
            audit_service=audit_service,
        )

        payload = service.list_workspace(
            auth_context=_build_auth_context(
                role="accountant",
                subject_id=seeded["accountant_subject_id"],
            ),
            request=_build_request("/api/v1/accounting/workspace"),
            search=None,
            limit=20,
            offset=0,
        )

    assert payload.total == 2
    assert [item.last_name for item in payload.items] == ["Adams", "Baker"]
    assert payload.items[0].accountant_task_total == 2
    assert payload.items[0].accountant_task_pending == 1
    assert payload.items[0].accountant_task_completed == 1
    assert payload.items[0].accountant_task_overdue == 1
    assert payload.items[1].accountant_task_total == 1
    assert payload.items[1].accountant_task_in_progress == 1
    assert all(item.last_name != "Clark" for item in payload.items)
    assert audit_service.events[0]["action"] == "accounting_workspace:read"


def test_accounting_workspace_export_helpers_share_same_columns_and_values() -> None:
    """Verify CSV and XLSX exports use the same header order and row values."""
    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    Base.metadata.create_all(engine)
    audit_service = _AuditServiceStub()
    with Session(engine) as session:
        seeded = _seed_accounting_workspace(session)
        service = AccountingWorkspaceService(
            workspace_dao=AccountingWorkspaceDAO(session=session),
            task_dao=OnboardingTaskDAO(session=session),
            audit_service=audit_service,
        )
        rows = service.list_workspace(
            auth_context=_build_auth_context(
                role="accountant",
                subject_id=seeded["accountant_subject_id"],
            ),
            request=_build_request("/api/v1/accounting/workspace"),
            search="payroll",
            limit=20,
            offset=0,
        ).items
        csv_payload = service.export_workspace(
            auth_context=_build_auth_context(
                role="accountant",
                subject_id=seeded["accountant_subject_id"],
            ),
            request=_build_request("/api/v1/accounting/workspace/export"),
            search="payroll",
            export_format="csv",
        )
        xlsx_payload = service.export_workspace(
            auth_context=_build_auth_context(
                role="accountant",
                subject_id=seeded["accountant_subject_id"],
            ),
            request=_build_request("/api/v1/accounting/workspace/export"),
            search="payroll",
            export_format="xlsx",
        )

    assert len(rows) == 1
    csv_lines = csv_payload.content.decode("utf-8").strip().splitlines()
    assert csv_lines[0].split(",") == list(ACCOUNTING_WORKSPACE_EXPORT_COLUMNS)
    workbook = load_workbook(filename=BytesIO(xlsx_payload.content))
    worksheet = workbook["accounting_workspace"]
    xlsx_values = list(worksheet.iter_rows(values_only=True))
    assert list(xlsx_values[0]) == list(ACCOUNTING_WORKSPACE_EXPORT_COLUMNS)
    assert str(xlsx_values[1][0]) == str(rows[0].onboarding_id)
    assert str(xlsx_values[1][1]) == str(rows[0].employee_id)
    assert xlsx_values[1][2] == rows[0].first_name
    assert xlsx_values[1][3] == rows[0].last_name
    assert xlsx_values[1][4] == rows[0].email
    assert workbook.sheetnames == ["accounting_workspace"]
