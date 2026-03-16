"""Integration tests for KPI snapshot reporting APIs."""

from __future__ import annotations

import csv
from datetime import UTC, datetime
from io import BytesIO, StringIO
from pathlib import Path
from uuid import uuid4

import pytest
from httpx import ASGITransport, AsyncClient
from openpyxl import load_workbook
from sqlalchemy import create_engine, select
from sqlalchemy.orm import Session

from hrm_backend.audit.models.event import AuditEvent
from hrm_backend.auth.dependencies.auth import get_current_auth_context
from hrm_backend.auth.schemas.token_claims import AuthContext
from hrm_backend.candidates.models.profile import CandidateProfile
from hrm_backend.core.models.base import Base
from hrm_backend.employee.models.hire_conversion import HireConversion
from hrm_backend.employee.models.onboarding import OnboardingRun, OnboardingTask
from hrm_backend.employee.models.profile import EmployeeProfile
from hrm_backend.interviews.models.interview import Interview
from hrm_backend.main import app
from hrm_backend.settings import AppSettings, get_settings
from hrm_backend.vacancies.models.offer import Offer
from hrm_backend.vacancies.models.pipeline_transition import PipelineTransition
from hrm_backend.vacancies.models.vacancy import Vacancy

pytestmark = pytest.mark.anyio


@pytest.fixture()
def sqlite_database_url(tmp_path: Path) -> str:
    """Provide temporary SQLite URL for KPI snapshot integration tests."""
    return f"sqlite+pysqlite:///{tmp_path / 'kpi_snapshots_api.db'}"


@pytest.fixture()
def configured_app(sqlite_database_url: str):
    """Configure app dependency overrides for KPI snapshot API tests."""
    settings = AppSettings(
        database_url=sqlite_database_url,
        redis_url="redis://localhost:6379/15",
        jwt_secret="integration-secret-with-minimum-32-bytes",
    )
    context_holder = {
        "context": AuthContext(
            subject_id=uuid4(),
            role="admin",
            session_id=uuid4(),
            token_id=uuid4(),
            expires_at=9999999999,
        )
    }

    def _get_settings_override() -> AppSettings:
        return settings

    def _get_auth_context_override() -> AuthContext:
        return context_holder["context"]

    app.dependency_overrides[get_settings] = _get_settings_override
    app.dependency_overrides[get_current_auth_context] = _get_auth_context_override

    engine = create_engine(sqlite_database_url, future=True)
    Base.metadata.create_all(engine)
    try:
        yield app, context_holder, engine
    finally:
        app.dependency_overrides.pop(get_settings, None)
        app.dependency_overrides.pop(get_current_auth_context, None)
        engine.dispose()


@pytest.fixture()
async def api_client(configured_app) -> AsyncClient:
    """Provide async API client for KPI snapshot integration tests."""
    configured, *_ = configured_app
    async with AsyncClient(
        transport=ASGITransport(app=configured),
        base_url="http://testserver",
    ) as client:
        yield client


def _load_audit_events(engine) -> list[AuditEvent]:
    """Load ordered audit events for export audit assertions."""
    with Session(engine) as session:
        statement = select(AuditEvent).order_by(
            AuditEvent.occurred_at,
            AuditEvent.event_id,
        )
        return list(session.execute(statement).scalars().all())


def _seed_kpi_sources(engine) -> None:
    """Insert one row per KPI source inside March 2026."""
    vacancy_id = "aaaaaaaa-aaaa-4aaa-8aaa-aaaaaaaaaaaa"
    candidate_id = "bbbbbbbb-bbbb-4bbb-8bbb-bbbbbbbbbbbb"
    offer_id = "cccccccc-cccc-4ccc-8ccc-cccccccccccc"
    transition_applied_id = "dddddddd-dddd-4ddd-8ddd-dddddddddddd"
    transition_hired_id = "eeeeeeee-eeee-4eee-8eee-eeeeeeeeeeee"
    conversion_id = "ffffffff-ffff-4fff-8fff-ffffffffffff"
    employee_id = "11111111-2222-4333-8444-555555555555"
    onboarding_id = "66666666-7777-4888-8999-000000000000"
    task_id = "12121212-3434-4567-8999-121212121212"

    with Session(engine) as session:
        session.add(
            Vacancy(
                vacancy_id=vacancy_id,
                title="Platform Engineer",
                description="Build platform foundations.",
                department="Engineering",
                status="open",
                created_at=datetime(2026, 3, 5, 9, 0, tzinfo=UTC),
                updated_at=datetime(2026, 3, 5, 9, 0, tzinfo=UTC),
            )
        )
        session.add(
            CandidateProfile(
                candidate_id=candidate_id,
                owner_subject_id="public",
                first_name="Ada",
                last_name="Lovelace",
                email="ada@example.com",
                phone=None,
                location=None,
                current_title=None,
                extra_data={},
                created_at=datetime(2026, 3, 6, 9, 0, tzinfo=UTC),
                updated_at=datetime(2026, 3, 6, 9, 0, tzinfo=UTC),
            )
        )
        session.add(
            PipelineTransition(
                transition_id=transition_applied_id,
                vacancy_id=vacancy_id,
                candidate_id=candidate_id,
                from_stage=None,
                to_stage="applied",
                reason="public_application",
                changed_by_sub="public",
                changed_by_role="public",
                transitioned_at=datetime(2026, 3, 6, 10, 0, tzinfo=UTC),
            )
        )
        session.add(
            Interview(
                interview_id="99999999-aaaa-4bbb-8ccc-999999999999",
                vacancy_id=vacancy_id,
                candidate_id=candidate_id,
                status="scheduled",
                calendar_sync_status="pending",
                schedule_version=1,
                scheduled_start_at=datetime(2026, 3, 10, 10, 0, tzinfo=UTC),
                scheduled_end_at=datetime(2026, 3, 10, 11, 0, tzinfo=UTC),
                timezone="UTC",
                location_kind="video",
                location_details=None,
                interviewer_staff_ids_json=["22222222-2222-4222-8222-222222222222"],
                calendar_event_id=None,
                candidate_token_nonce=None,
                candidate_token_hash=None,
                candidate_token_expires_at=None,
                candidate_response_status="pending",
                candidate_response_note=None,
                cancelled_by=None,
                cancel_reason_code=None,
                calendar_sync_reason_code=None,
                calendar_sync_error_detail=None,
                created_by_staff_id="33333333-3333-4333-8333-333333333333",
                updated_by_staff_id="33333333-3333-4333-8333-333333333333",
                created_at=datetime(2026, 3, 7, 9, 0, tzinfo=UTC),
                updated_at=datetime(2026, 3, 7, 9, 0, tzinfo=UTC),
                last_synced_at=None,
            )
        )
        session.add(
            Offer(
                offer_id=offer_id,
                vacancy_id=vacancy_id,
                candidate_id=candidate_id,
                status="accepted",
                terms_summary="Base offer",
                proposed_start_date=None,
                expires_at=None,
                note=None,
                sent_at=datetime(2026, 3, 12, 9, 0, tzinfo=UTC),
                sent_by_staff_id="44444444-4444-4444-8444-444444444444",
                decision_at=datetime(2026, 3, 15, 9, 0, tzinfo=UTC),
                decision_note=None,
                decision_recorded_by_staff_id="44444444-4444-4444-8444-444444444444",
                created_at=datetime(2026, 3, 11, 9, 0, tzinfo=UTC),
                updated_at=datetime(2026, 3, 15, 9, 0, tzinfo=UTC),
            )
        )
        session.add(
            PipelineTransition(
                transition_id=transition_hired_id,
                vacancy_id=vacancy_id,
                candidate_id=candidate_id,
                from_stage="offer",
                to_stage="hired",
                reason=None,
                changed_by_sub="44444444-4444-4444-8444-444444444444",
                changed_by_role="hr",
                transitioned_at=datetime(2026, 3, 16, 9, 0, tzinfo=UTC),
            )
        )
        session.add(
            HireConversion(
                conversion_id=conversion_id,
                vacancy_id=vacancy_id,
                candidate_id=candidate_id,
                offer_id=offer_id,
                hired_transition_id=transition_hired_id,
                status="ready",
                candidate_snapshot_json={"first_name": "Ada", "last_name": "Lovelace"},
                offer_snapshot_json={"status": "accepted"},
                converted_at=datetime(2026, 3, 16, 10, 0, tzinfo=UTC),
                converted_by_staff_id="44444444-4444-4444-8444-444444444444",
            )
        )
        session.add(
            EmployeeProfile(
                employee_id=employee_id,
                hire_conversion_id=conversion_id,
                vacancy_id=vacancy_id,
                candidate_id=candidate_id,
                first_name="Ada",
                last_name="Lovelace",
                email="ada@example.com",
                phone=None,
                location=None,
                current_title=None,
                extra_data_json={},
                offer_terms_summary="Base offer",
                start_date=None,
                staff_account_id=None,
                created_by_staff_id="44444444-4444-4444-8444-444444444444",
                created_at=datetime(2026, 3, 16, 12, 0, tzinfo=UTC),
                updated_at=datetime(2026, 3, 16, 12, 0, tzinfo=UTC),
            )
        )
        session.add(
            OnboardingRun(
                onboarding_id=onboarding_id,
                employee_id=employee_id,
                hire_conversion_id=conversion_id,
                status="started",
                started_at=datetime(2026, 3, 17, 9, 0, tzinfo=UTC),
                started_by_staff_id="44444444-4444-4444-8444-444444444444",
            )
        )
        session.add(
            OnboardingTask(
                task_id=task_id,
                onboarding_id=onboarding_id,
                template_id="22222222-3333-4444-8555-666666666666",
                template_item_id="77777777-8888-4999-8aaa-bbbbbbbbbbbb",
                code="welcome",
                title="Welcome",
                description=None,
                sort_order=10,
                is_required=True,
                status="completed",
                assigned_role=None,
                assigned_staff_id=None,
                due_at=None,
                completed_at=datetime(2026, 3, 18, 9, 0, tzinfo=UTC),
                created_at=datetime(2026, 3, 17, 9, 0, tzinfo=UTC),
                updated_at=datetime(2026, 3, 18, 9, 0, tzinfo=UTC),
            )
        )
        session.commit()


async def test_admin_can_rebuild_and_read_kpi_snapshots(
    configured_app,
    api_client: AsyncClient,
) -> None:
    """Verify admin can rebuild and read KPI snapshot rows."""
    _, _, engine = configured_app
    _seed_kpi_sources(engine)

    rebuild_response = await api_client.post(
        "/api/v1/reporting/kpi-snapshots/rebuild",
        json={"period_month": "2026-03-01"},
    )
    assert rebuild_response.status_code == 200
    payload = rebuild_response.json()
    assert payload["period_month"] == "2026-03-01"
    assert len(payload["metrics"]) == 8

    read_response = await api_client.get(
        "/api/v1/reporting/kpi-snapshots",
        params={"period_month": "2026-03-01"},
    )
    assert read_response.status_code == 200
    read_payload = read_response.json()
    assert len(read_payload["metrics"]) == 8
    vacancy_metric = next(
        item for item in read_payload["metrics"] if item["metric_key"] == "vacancies_created_count"
    )
    assert vacancy_metric["metric_value"] == 1


async def test_kpi_snapshot_read_returns_empty_payload_when_missing(
    api_client: AsyncClient,
) -> None:
    """Verify read API returns empty metrics when snapshot month is missing."""
    read_response = await api_client.get(
        "/api/v1/reporting/kpi-snapshots",
        params={"period_month": "2026-02-01"},
    )
    assert read_response.status_code == 200
    payload = read_response.json()
    assert payload["metrics"] == []


async def test_kpi_snapshot_read_does_not_fallback_to_live_aggregation(
    configured_app,
    api_client: AsyncClient,
) -> None:
    """Verify read API returns empty payload even when source data exists."""
    _, context_holder, engine = configured_app
    _seed_kpi_sources(engine)
    context_holder["context"] = AuthContext(
        subject_id=uuid4(),
        role="leader",
        session_id=uuid4(),
        token_id=uuid4(),
        expires_at=9999999999,
    )

    read_response = await api_client.get(
        "/api/v1/reporting/kpi-snapshots",
        params={"period_month": "2026-03-01"},
    )
    assert read_response.status_code == 200
    payload = read_response.json()
    assert payload["metrics"] == []


async def test_kpi_snapshot_rbac_allows_leader_reads_and_denies_rebuild(
    configured_app,
    api_client: AsyncClient,
) -> None:
    """Verify leader can read KPI snapshots but cannot rebuild them."""
    _, context_holder, _ = configured_app
    context_holder["context"] = AuthContext(
        subject_id=uuid4(),
        role="leader",
        session_id=uuid4(),
        token_id=uuid4(),
        expires_at=9999999999,
    )

    read_response = await api_client.get(
        "/api/v1/reporting/kpi-snapshots",
        params={"period_month": "2026-03-01"},
    )
    assert read_response.status_code == 200

    rebuild_response = await api_client.post(
        "/api/v1/reporting/kpi-snapshots/rebuild",
        json={"period_month": "2026-03-01"},
    )
    assert rebuild_response.status_code == 403


@pytest.mark.parametrize("role", ["hr", "manager", "employee", "accountant"])
async def test_kpi_snapshot_read_denies_non_privileged_roles(
    configured_app,
    api_client: AsyncClient,
    role: str,
) -> None:
    """Verify non-privileged roles are denied KPI snapshot reads."""
    _, context_holder, _ = configured_app
    context_holder["context"] = AuthContext(
        subject_id=uuid4(),
        role=role,  # type: ignore[arg-type]
        session_id=uuid4(),
        token_id=uuid4(),
        expires_at=9999999999,
    )

    read_response = await api_client.get(
        "/api/v1/reporting/kpi-snapshots",
        params={"period_month": "2026-03-01"},
    )
    assert read_response.status_code == 403


async def test_kpi_snapshot_invalid_month_returns_422(
    api_client: AsyncClient,
) -> None:
    """Verify invalid period month values return 422 responses."""
    rebuild_response = await api_client.post(
        "/api/v1/reporting/kpi-snapshots/rebuild",
        json={"period_month": "2026-03-15"},
    )
    assert rebuild_response.status_code == 422

    read_response = await api_client.get(
        "/api/v1/reporting/kpi-snapshots",
        params={"period_month": "2026-03-15"},
    )
    assert read_response.status_code == 422


async def test_admin_can_export_kpi_snapshot_as_csv_and_is_audited(
    configured_app,
    api_client: AsyncClient,
) -> None:
    """Verify admin can export stored KPI snapshot as CSV attachment and audit is recorded."""
    _, _, engine = configured_app
    _seed_kpi_sources(engine)

    rebuild_response = await api_client.post(
        "/api/v1/reporting/kpi-snapshots/rebuild",
        json={"period_month": "2026-03-01"},
    )
    assert rebuild_response.status_code == 200

    export_response = await api_client.get(
        "/api/v1/reporting/kpi-snapshots/export",
        params={"period_month": "2026-03-01", "format": "csv"},
    )
    assert export_response.status_code == 200
    assert export_response.headers["content-type"].startswith("text/csv")
    assert export_response.headers["content-disposition"].startswith(
        'attachment; filename="kpi-snapshot-2026-03-01-'
    )
    assert export_response.headers["content-disposition"].endswith('.csv"')

    rows = list(csv.reader(StringIO(export_response.text)))
    header = rows[0]
    metric_key_index = header.index("metric_key")
    metric_value_index = header.index("metric_value")
    vacancy_row = next(
        row
        for row in rows[1:]
        if row[metric_key_index] == "vacancies_created_count"
    )
    assert vacancy_row[metric_value_index] == "1"

    events = _load_audit_events(engine)
    assert any(
        event.action == "kpi_snapshot:export"
        and event.result == "success"
        and event.resource_id == "2026-03-01"
        for event in events
    )


async def test_admin_can_export_kpi_snapshot_as_xlsx(
    configured_app,
    api_client: AsyncClient,
) -> None:
    """Verify KPI snapshot export supports XLSX attachment rendering."""
    _, _, engine = configured_app
    _seed_kpi_sources(engine)

    rebuild_response = await api_client.post(
        "/api/v1/reporting/kpi-snapshots/rebuild",
        json={"period_month": "2026-03-01"},
    )
    assert rebuild_response.status_code == 200

    export_response = await api_client.get(
        "/api/v1/reporting/kpi-snapshots/export",
        params={"period_month": "2026-03-01", "format": "xlsx"},
    )
    assert export_response.status_code == 200
    assert export_response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    workbook = load_workbook(filename=BytesIO(export_response.content))
    worksheet = workbook["kpi_snapshot"]
    values = list(worksheet.iter_rows(values_only=True))
    assert values[0][0] == "period_month"
    assert values[1][0] == "2026-03-01"
    assert values[1][1] == "vacancies_created_count"


async def test_kpi_snapshot_export_allows_leader_reads(
    configured_app,
    api_client: AsyncClient,
) -> None:
    """Verify leader can export stored KPI snapshots but cannot rebuild them."""
    _, context_holder, engine = configured_app
    _seed_kpi_sources(engine)

    rebuild_response = await api_client.post(
        "/api/v1/reporting/kpi-snapshots/rebuild",
        json={"period_month": "2026-03-01"},
    )
    assert rebuild_response.status_code == 200

    context_holder["context"] = AuthContext(
        subject_id=uuid4(),
        role="leader",
        session_id=uuid4(),
        token_id=uuid4(),
        expires_at=9999999999,
    )

    export_response = await api_client.get(
        "/api/v1/reporting/kpi-snapshots/export",
        params={"period_month": "2026-03-01", "format": "csv"},
    )
    assert export_response.status_code == 200
    assert "vacancies_created_count" in export_response.text


@pytest.mark.parametrize("role", ["hr", "manager", "employee", "accountant"])
async def test_kpi_snapshot_export_denies_non_privileged_roles(
    configured_app,
    api_client: AsyncClient,
    role: str,
) -> None:
    """Verify non-privileged roles are denied KPI snapshot exports."""
    _, context_holder, _ = configured_app
    context_holder["context"] = AuthContext(
        subject_id=uuid4(),
        role=role,  # type: ignore[arg-type]
        session_id=uuid4(),
        token_id=uuid4(),
        expires_at=9999999999,
    )

    export_response = await api_client.get(
        "/api/v1/reporting/kpi-snapshots/export",
        params={"period_month": "2026-03-01", "format": "csv"},
    )
    assert export_response.status_code == 403


async def test_kpi_snapshot_export_invalid_month_returns_422(
    api_client: AsyncClient,
) -> None:
    """Verify invalid export period month values return 422 responses."""
    export_response = await api_client.get(
        "/api/v1/reporting/kpi-snapshots/export",
        params={"period_month": "2026-03-15", "format": "csv"},
    )
    assert export_response.status_code == 422
