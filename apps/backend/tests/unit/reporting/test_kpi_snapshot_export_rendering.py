"""Unit tests for KPI snapshot export rendering helpers."""

from __future__ import annotations

import csv
from datetime import UTC, date, datetime
from io import BytesIO, StringIO

from openpyxl import load_workbook

from hrm_backend.reporting.schemas.kpi_snapshot import KpiSnapshotMetric, KpiSnapshotReadResponse
from hrm_backend.reporting.utils.exports import (
    KPI_SNAPSHOT_EXPORT_COLUMNS,
    render_kpi_snapshot_csv,
    render_kpi_snapshot_xlsx,
)


def test_render_kpi_snapshot_csv_has_stable_header_and_values() -> None:
    """Verify KPI snapshot CSV export keeps stable columns and serializes values."""
    snapshot = KpiSnapshotReadResponse(
        period_month=date(2026, 3, 1),
        metrics=[
            KpiSnapshotMetric(
                metric_key="vacancies_created_count",
                metric_value=1,
                generated_at=datetime(2026, 3, 16, 12, 0, tzinfo=UTC),
            )
        ],
    )

    payload = render_kpi_snapshot_csv(snapshot)
    rows = list(csv.reader(StringIO(payload.decode("utf-8"))))

    assert tuple(rows[0]) == KPI_SNAPSHOT_EXPORT_COLUMNS
    assert rows[1][0] == "2026-03-01"
    assert rows[1][1] == "vacancies_created_count"
    assert rows[1][2] == "1"
    assert rows[1][3].startswith("2026-03-16T12:00:00")


def test_render_kpi_snapshot_xlsx_has_stable_sheet_and_values() -> None:
    """Verify KPI snapshot XLSX export uses canonical sheet title and row shape."""
    snapshot = KpiSnapshotReadResponse(
        period_month=date(2026, 3, 1),
        metrics=[
            KpiSnapshotMetric(
                metric_key="offers_sent_count",
                metric_value=2,
                generated_at=None,
            )
        ],
    )

    payload = render_kpi_snapshot_xlsx(snapshot)
    workbook = load_workbook(filename=BytesIO(payload))
    worksheet = workbook["kpi_snapshot"]
    values = list(worksheet.iter_rows(values_only=True))

    assert tuple(values[0]) == KPI_SNAPSHOT_EXPORT_COLUMNS
    assert values[1][0] == "2026-03-01"
    assert values[1][1] == "offers_sent_count"
    assert values[1][2] == 2
    assert values[1][3] is None

